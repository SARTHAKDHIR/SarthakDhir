"""
Smart Vendor Order Allocation Based on Capacity v1.1
Professional Supply Chain Demand Allocation Tool
- Rollover: unallocated qty from Month N is added to Month N+1 demand first
- Light professional colour scheme
"""

import sys
import subprocess
import os

def _ensure(pkg):
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", pkg,
             "--quiet", "--disable-pip-version-check"],
            **({'creationflags': 0x08000000} if sys.platform == "win32" else {})
        )

_ensure("pandas")
_ensure("openpyxl")

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import threading

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = dict(
    bg        = "#F8FAFC",
    panel     = "#FFFFFF",
    card      = "#FFFFFF",
    card2     = "#CCFBF1",
    accent    = "#0D9488",
    acch      = "#0F766E",
    success   = "#059669",
    warn      = "#D97706",
    error     = "#DC2626",
    text      = "#111928",
    dim       = "#6B7280",
    border    = "#D1D5DB",
    entry     = "#F9FAFB",
    sel       = "#CCFBF1",
    sel_text  = "#134E4A",
    step_act  = "#0D9488",
    step_don  = "#059669",
    step_pnd  = "#E5E7EB",
    step_pfg  = "#6B7280",
    row_even  = "#F9FAFB",
    row_odd   = "#FFFFFF",
    hdr       = "#CCFBF1",
    hdr_text  = "#0D9488",
    wbg       = "#FEF3C7",
    wfg       = "#92400E",
)

FF = "Segoe UI" if os.name == "nt" else "Helvetica"

def fnt(size=10, bold=False, italic=False):
    return (FF, size, "bold" if bold else "normal", "italic" if italic else "roman")

class ScrollFrame(tk.Frame):
    def __init__(self, parent, **kw):
        bg = kw.pop("bg", C["bg"])
        super().__init__(parent, bg=bg, **kw)
        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self.canvas, bg=bg)
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_inner)
        self.canvas.bind("<Configure>", self._on_canvas)
        for w in (self, self.inner, self.canvas):
            w.bind("<MouseWheel>", self._scroll)
            w.bind("<Button-4>",   self._scroll)
            w.bind("<Button-5>",   self._scroll)

    def _on_inner(self, _=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas(self, e):
        self.canvas.itemconfig(self._win, width=e.width)

    def _scroll(self, e):
        if e.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif e.num == 5:
            self.canvas.yview_scroll(1, "units")
        else:
            self.canvas.yview_scroll(int(-e.delta / 120), "units")


def make_btn(parent, text, cmd=None, kind="primary", width=None):
    colours = {
        "primary": (C["accent"],  C["acch"],   "#FFFFFF"),
        "success": (C["success"], "#065F46",   "#FFFFFF"),
        "danger":  (C["error"],   "#9B1C1C",   "#FFFFFF"),
        "ghost":   ("#F3F4F6",    "#E5E7EB",   C["text"]),
        "dim":     ("#E5E7EB",    "#D1D5DB",   C["dim"]),
    }
    bg, hov, fg = colours.get(kind, colours["primary"])
    kw = {"width": width} if width else {}
    btn = tk.Button(
        parent, text=text, command=cmd,
        bg=bg, fg=fg, activebackground=hov, activeforeground=fg,
        font=fnt(10, bold=(kind in ("primary", "success"))),
        relief="flat", cursor="hand2", bd=0, padx=14, pady=6, **kw
    )
    btn.bind("<Enter>", lambda e: btn.config(bg=hov))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn


def make_combo(parent, values=(), textvariable=None, width=28):
    st = ttk.Style()
    st.theme_use("default")
    st.configure("Pro.TCombobox",
        fieldbackground=C["entry"], background=C["panel"],
        foreground=C["text"], arrowcolor=C["accent"],
        bordercolor=C["border"], lightcolor=C["border"], darkcolor=C["border"],
        selectbackground=C["sel"], selectforeground=C["sel_text"])
    st.map("Pro.TCombobox",
        fieldbackground=[("readonly", C["entry"])],
        foreground=[("readonly", C["text"])])
    return ttk.Combobox(parent, values=list(values), textvariable=textvariable,
                        width=width, style="Pro.TCombobox", state="readonly")


def hsep(parent, pad=8):
    tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", pady=pad)


def card(parent, **kw):
    return tk.Frame(parent, bg=C["card"],
                    highlightbackground=C["border"], highlightthickness=1, **kw)


def slabel(parent, text):
    tk.Label(parent, text=text, bg=C["card"], fg=C["accent"],
             font=fnt(11, bold=True)).pack(anchor="w", padx=16, pady=(14, 3))


def hlabel(parent, text):
    tk.Label(parent, text=text, bg=C["card"], fg=C["dim"],
             font=fnt(9), wraplength=720, justify="left").pack(anchor="w", padx=16, pady=(0, 6))


class DragList(tk.Frame):
    def __init__(self, parent, width=26, height=7, **kw):
        super().__init__(parent, bg=C["card"], **kw)
        self._drag_idx = None
        sb = tk.Scrollbar(self, orient="vertical")
        self.lb = tk.Listbox(self, yscrollcommand=sb.set,
            bg=C["entry"], fg=C["text"],
            selectbackground=C["sel"], selectforeground=C["sel_text"],
            font=fnt(10), relief="flat", bd=0,
            activestyle="none", width=width, height=height,
            highlightthickness=1, highlightbackground=C["border"],
            highlightcolor=C["accent"])
        sb.config(command=self.lb.yview)
        sb.pack(side="right", fill="y")
        self.lb.pack(side="left", fill="both", expand=True)
        self.lb.bind("<ButtonPress-1>",   self._start)
        self.lb.bind("<B1-Motion>",       self._drag)
        self.lb.bind("<ButtonRelease-1>", self._end)
        bb = tk.Frame(self, bg=C["card"])
        bb.pack(side="bottom", fill="x", pady=(4, 0))
        make_btn(bb, "Up",   self._up,   kind="ghost",  width=8).pack(side="left",  padx=2, pady=2)
        make_btn(bb, "Down", self._down, kind="ghost",  width=8).pack(side="left",  padx=2, pady=2)
        make_btn(bb, "Del",  self._del,  kind="danger", width=8).pack(side="right", padx=2, pady=2)

    def items(self):
        return list(self.lb.get(0, tk.END))

    def add(self, item):
        if item and item not in self.items():
            self.lb.insert(tk.END, item)

    def clear(self):
        self.lb.delete(0, tk.END)

    def _start(self, e):
        self._drag_idx = self.lb.nearest(e.y)

    def _drag(self, e):
        idx = self.lb.nearest(e.y)
        if idx == self._drag_idx:
            return
        val = self.lb.get(self._drag_idx)
        self.lb.delete(self._drag_idx)
        self.lb.insert(idx, val)
        self.lb.selection_set(idx)
        self._drag_idx = idx

    def _end(self, _):
        self._drag_idx = None

    def _up(self):
        sel = self.lb.curselection()
        if not sel or sel[0] == 0:
            return
        i, v = sel[0], self.lb.get(sel[0])
        self.lb.delete(i); self.lb.insert(i - 1, v); self.lb.selection_set(i - 1)

    def _down(self):
        sel = self.lb.curselection()
        if not sel or sel[0] == self.lb.size() - 1:
            return
        i, v = sel[0], self.lb.get(sel[0])
        self.lb.delete(i); self.lb.insert(i + 1, v); self.lb.selection_set(i + 1)

    def _del(self):
        sel = self.lb.curselection()
        if sel:
            self.lb.delete(sel[0])


class ColChooser(tk.Frame):
    def __init__(self, parent, title_r="Selected", multi=True, height=6, **kw):
        super().__init__(parent, bg=C["card"], **kw)
        self._multi = multi
        self._all_cols = []

        sf = tk.Frame(self, bg=C["card"])
        sf.pack(fill="x", padx=4, pady=(4, 2))
        tk.Label(sf, text="Search:", bg=C["card"], fg=C["dim"], font=fnt(9)).pack(side="left")
        self._svar = tk.StringVar()
        self._svar.trace_add("write", self._filter)
        tk.Entry(sf, textvariable=self._svar,
            bg=C["entry"], fg=C["text"], insertbackground=C["text"],
            relief="flat", font=fnt(10),
            highlightthickness=1, highlightbackground=C["border"],
            highlightcolor=C["accent"]).pack(side="left", fill="x", expand=True, padx=4)

        body = tk.Frame(self, bg=C["card"])
        body.pack(fill="both", expand=True, padx=4, pady=2)

        self._af = self._make_lb(body, "Available", height)
        self._af.pack(side="left", fill="both", expand=True)

        mid = tk.Frame(body, bg=C["card"])
        mid.pack(side="left", padx=6)
        make_btn(mid, "->",  self._add,     kind="primary", width=3).pack(pady=3)
        make_btn(mid, "<-",  self._remove,  kind="ghost",   width=3).pack(pady=3)
        make_btn(mid, ">>",  self._add_all, kind="dim",     width=3).pack(pady=10)
        make_btn(mid, "<<",  self._clr,     kind="danger",  width=3).pack(pady=3)

        right = tk.Frame(body, bg=C["card"])
        right.pack(side="left", fill="both", expand=True)
        tk.Label(right, text=title_r, bg=C["card"], fg=C["dim"],
                 font=fnt(9, bold=True)).pack(anchor="w")
        if multi:
            self._sel = DragList(right, height=height)
            self._sel.pack(fill="both", expand=True)
        else:
            self._sf = self._make_lb(right, "", height)
            self._sf.pack(fill="both", expand=True)

        self._af.lb.bind("<Double-1>", lambda e: self._add())

    def set_columns(self, cols):
        self._all_cols = list(cols)
        self._refresh()

    def get_selected(self):
        if self._multi:
            return self._sel.items()
        return list(self._sf.lb.get(0, tk.END))

    def set_prefix_filter(self, prefix):
        self._svar.set(prefix)

    def clear_selected(self):
        if self._multi:
            self._sel.clear()
        else:
            self._sf.lb.delete(0, tk.END)

    def _make_lb(self, parent, lbl, h):
        f = tk.Frame(parent, bg=C["card"])
        if lbl:
            tk.Label(f, text=lbl, bg=C["card"], fg=C["dim"],
                     font=fnt(9, bold=True)).pack(anchor="w")
        sb = tk.Scrollbar(f, orient="vertical")
        lb = tk.Listbox(f, yscrollcommand=sb.set,
            bg=C["entry"], fg=C["text"],
            selectbackground=C["sel"], selectforeground=C["sel_text"],
            font=fnt(10), relief="flat", bd=0,
            highlightthickness=1, highlightbackground=C["border"],
            highlightcolor=C["accent"],
            activestyle="none", height=h, selectmode=tk.EXTENDED)
        sb.config(command=lb.yview)
        sb.pack(side="right", fill="y")
        lb.pack(side="left", fill="both", expand=True)
        lb.bind("<MouseWheel>", lambda e: lb.yview_scroll(int(-e.delta / 120), "units"))
        lb.bind("<Button-4>",   lambda e: lb.yview_scroll(-1, "units"))
        lb.bind("<Button-5>",   lambda e: lb.yview_scroll(1, "units"))
        f.lb = lb
        return f

    def _refresh(self):
        term = self._svar.get().lower()
        self._af.lb.delete(0, tk.END)
        for c in self._all_cols:
            if term in str(c).lower():
                self._af.lb.insert(tk.END, c)

    def _filter(self, *_):
        self._refresh()

    def _add(self):
        for i in self._af.lb.curselection():
            v = self._af.lb.get(i)
            if self._multi:
                self._sel.add(v)
            else:
                if v not in self._sf.lb.get(0, tk.END):
                    self._sf.lb.insert(tk.END, v)

    def _remove(self):
        if self._multi:
            for i in reversed(self._sel.lb.curselection()):
                self._sel.lb.delete(i)
        else:
            for i in reversed(self._sf.lb.curselection()):
                self._sf.lb.delete(i)

    def _add_all(self):
        for c in self._all_cols:
            if self._multi:
                self._sel.add(c)
            else:
                if c not in self._sf.lb.get(0, tk.END):
                    self._sf.lb.insert(tk.END, c)

    def _clr(self):
        if self._multi:
            self._sel.clear()
        else:
            self._sf.lb.delete(0, tk.END)


class MonthCapTable(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=C["card"], **kw)
        self._rows = []
        self._cols = []

        hdr = tk.Frame(self, bg=C["card2"])
        hdr.pack(fill="x", padx=2, pady=(4, 0))
        tk.Label(hdr, text="Month Label", bg=C["card2"], fg=C["dim"],
                 font=fnt(9, bold=True), width=22, anchor="w").pack(side="left", padx=4, pady=4)
        tk.Label(hdr, text="Capacity Column in Cap Sheet", bg=C["card2"], fg=C["dim"],
                 font=fnt(9, bold=True), width=30, anchor="w").pack(side="left", padx=4)

        self._body = tk.Frame(self, bg=C["card"])
        self._body.pack(fill="x", padx=2)

        make_btn(self, "+ Add Month", self.add_row, kind="success", width=16).pack(
            anchor="w", padx=4, pady=8)

    def set_columns(self, cols):
        self._cols = list(cols)
        for _, cv, rf in self._rows:
            for w in rf.winfo_children():
                if isinstance(w, ttk.Combobox):
                    w["values"] = self._cols

    def get_mapping(self):
        result = []
        for mv, cv, _ in self._rows:
            m = mv.get().strip()
            c = cv.get().strip()
            if m and c:
                result.append((m, c))
        return result

    def add_row(self, month="", col=""):
        rf = tk.Frame(self._body, bg=C["entry"],
                      highlightbackground=C["border"], highlightthickness=1)
        rf.pack(fill="x", pady=3, padx=2)
        mv = tk.StringVar(value=month)
        tk.Entry(rf, textvariable=mv, width=20,
            bg=C["entry"], fg=C["text"], insertbackground=C["text"],
            relief="flat", font=fnt(10), highlightthickness=0).pack(
                side="left", padx=(8, 8), pady=6)
        cv = tk.StringVar(value=col)
        cb = make_combo(rf, values=self._cols, textvariable=cv, width=28)
        cb.pack(side="left", padx=(0, 8))

        def remove():
            rf.destroy()
            self._rows[:] = [(a, b, c) for a, b, c in self._rows if c is not rf]

        make_btn(rf, "X", remove, kind="danger", width=3).pack(side="left", padx=(0, 6))
        self._rows.append((mv, cv, rf))

    def clear(self):
        for _, _, rf in self._rows:
            rf.destroy()
        self._rows.clear()


class StepBar(tk.Frame):
    def __init__(self, parent, steps, **kw):
        super().__init__(parent, bg=C["panel"], **kw)
        self._items = []
        for i, s in enumerate(steps):
            num = tk.Label(self, text=str(i + 1), width=3,
                           bg=C["step_pnd"], fg=C["step_pfg"],
                           font=fnt(10, bold=True), relief="flat")
            num.pack(side="left", ipady=6)
            lbl = tk.Label(self, text=s, bg=C["panel"], fg=C["dim"], font=fnt(10))
            lbl.pack(side="left", padx=(4, 0))
            if i < len(steps) - 1:
                tk.Label(self, text="  >  ", bg=C["panel"], fg=C["border"],
                         font=fnt(10)).pack(side="left")
            self._items.append((num, lbl))

    def set_step(self, active):
        for i, (n, l) in enumerate(self._items):
            if i < active:
                n.config(bg=C["step_don"], fg="#FFFFFF")
                l.config(fg=C["text"], font=fnt(10))
            elif i == active:
                n.config(bg=C["step_act"], fg="#FFFFFF")
                l.config(fg=C["accent"], font=fnt(10, bold=True))
            else:
                n.config(bg=C["step_pnd"], fg=C["step_pfg"])
                l.config(fg=C["dim"], font=fnt(10))


class ResultsTable(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=C["panel"], **kw)
        st = ttk.Style()
        st.theme_use("default")
        st.configure("Res.Treeview",
            background=C["row_odd"], foreground=C["text"],
            rowheight=26, fieldbackground=C["row_odd"],
            font=(FF, 9), borderwidth=0)
        st.configure("Res.Treeview.Heading",
            background=C["hdr"], foreground=C["hdr_text"],
            font=(FF, 9, "bold"), relief="flat", padding=6)
        st.map("Res.Treeview",
            background=[("selected", C["sel"])],
            foreground=[("selected", C["sel_text"])])
        st.layout("Res.Treeview", [("Treeview.treearea", {"sticky": "nswe"})])

        xsb = tk.Scrollbar(self, orient="horizontal")
        ysb = tk.Scrollbar(self, orient="vertical")
        self.tree = ttk.Treeview(self, style="Res.Treeview",
                                 yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        ysb.config(command=self.tree.yview)
        xsb.config(command=self.tree.xview)
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure("odd",      background=C["row_odd"])
        self.tree.tag_configure("even",     background=C["row_even"])
        self.tree.tag_configure("rollover", background="#F0FDFA", foreground=C["accent"])
        self.tree.bind("<MouseWheel>",
            lambda e: self.tree.yview_scroll(int(-e.delta / 120), "units"))
        self.tree.bind("<Button-4>", lambda e: self.tree.yview_scroll(-1, "units"))
        self.tree.bind("<Button-5>", lambda e: self.tree.yview_scroll(1, "units"))

    def load(self, df):
        self.tree.delete(*self.tree.get_children())
        cols = list(df.columns)
        self.tree["columns"] = cols
        self.tree["show"] = "headings"
        roll_idx = cols.index("Rolled to Next") if "Rolled to Next" in cols else (cols.index("Rollover In") if "Rollover In" in cols else None)
        for c in cols:
            w = max(90, min(240, len(str(c)) * 10))
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, minwidth=60, anchor="center")
        for i, (_, row) in enumerate(df.iterrows()):
            vals = ["" if pd.isna(v) else str(v) for v in row]
            has_roll = roll_idx is not None and str(row.iloc[roll_idx]) not in ("0", "", "nan", "0.0", "0")
            tag = "rollover" if has_roll else ("even" if i % 2 == 0 else "odd")
            self.tree.insert("", tk.END, values=vals, tags=(tag,))


class Toast:
    @staticmethod
    def show(root, msg, kind="success", duration=3000):
        cols = {"success": (C["success"], "#FFFFFF"),
                "error":   (C["error"],   "#FFFFFF"),
                "warn":    (C["warn"],    "#FFFFFF")}
        bg, fg = cols.get(kind, (C["accent"], "#FFFFFF"))
        t = tk.Toplevel(root)
        t.overrideredirect(True)
        t.attributes("-topmost", True)
        t.configure(bg=bg)
        tk.Label(t, text="  " + msg + "  ", bg=bg, fg=fg,
                 font=fnt(10, bold=True), padx=14, pady=10).pack()
        t.update_idletasks()
        rx = root.winfo_x() + root.winfo_width() // 2 - t.winfo_width() // 2
        ry = root.winfo_y() + root.winfo_height() - t.winfo_height() - 60
        t.geometry(f"+{rx}+{ry}")
        root.after(duration, t.destroy)


def run_allocation(demand_df, cap_df, item_col, month_cols,
                   vendor_cols, vendor_name_col,
                   month_cap_mapping, priority_col):
    """
    Allocation with rollover:
      - Each month's total demand = (unallocated from previous month) + (own demand)
      - Rolled-over qty is attempted first before fresh demand.
      - If the last month still has unallocated qty, a warning is raised.
    """
    months = [m for m, _ in month_cap_mapping]
    demand_month_map = {}
    for i, label in enumerate(months):
        if i < len(month_cols):
            demand_month_map[label] = month_cols[i]

    cap_dict = {}
    cdf = cap_df.copy()
    cdf.columns = cdf.columns.str.strip()

    for _, row in cdf.iterrows():
        vendor = str(row[vendor_name_col]).strip()
        if not vendor or vendor == "nan":
            continue
        prio = 999
        if priority_col and priority_col in cdf.columns:
            try:
                prio = int(row[priority_col])
            except Exception:
                prio = 999
        capacity = {}
        for label, cap_col in month_cap_mapping:
            try:
                capacity[label] = float(row[cap_col]) if pd.notna(row[cap_col]) else 0.0
            except Exception:
                capacity[label] = 0.0
        cap_dict[vendor] = {
            "priority": prio,
            "capacity": capacity,
            "used":     {m: 0.0 for m in months},
        }

    alloc_dict = {}
    warnings   = []

    for _, row in demand_df.iterrows():
        item = str(row[item_col]).strip()
        if not item or item == "nan":
            continue

        vendors = []
        for vc in vendor_cols:
            v = row.get(vc)
            if v is not None and pd.notna(v):
                vn = str(v).strip()
                if vn not in ("**", "", "nan") and vn not in vendors:
                    vendors.append(vn)
        valid = [v for v in vendors if v in cap_dict]

        carry = 0.0   # unallocated qty rolling in from previous month

        for month_idx, label in enumerate(months):
            dcol = demand_month_map.get(label)
            d_val = 0.0
            if dcol:
                raw = row.get(dcol, 0)
                d_val = float(raw) if not pd.isna(raw) else 0.0

            # Total demand = rollover from last month + own fresh demand
            total_demand = carry + d_val
            carry = 0.0

            if total_demand <= 0 or not valid:
                carry = total_demand   # carry forward if no vendors
                continue

            remaining = total_demand

            # Priority-sorted allocation pass
            for v in sorted(valid, key=lambda x: cap_dict[x]["priority"]):
                if remaining <= 0:
                    break
                avail = cap_dict[v]["capacity"][label] - cap_dict[v]["used"][label]
                if avail <= 0:
                    continue
                alloc = min(remaining, avail)
                cap_dict[v]["used"][label] += alloc
                remaining -= alloc
                key = (item, v)
                if key not in alloc_dict:
                    alloc_dict[key] = {m: 0.0 for m in months}
                alloc_dict[key][label] += alloc

            # Second pass: give remainder to most-available vendor
            while remaining > 0.1:
                candidates = [
                    (v, cap_dict[v]["capacity"][label] - cap_dict[v]["used"][label])
                    for v in valid
                    if cap_dict[v]["capacity"][label] - cap_dict[v]["used"][label] > 0
                ]
                if not candidates:
                    break
                v = max(candidates, key=lambda x: x[1])[0]
                avail = cap_dict[v]["capacity"][label] - cap_dict[v]["used"][label]
                alloc = min(remaining, avail)
                cap_dict[v]["used"][label] += alloc
                remaining -= alloc
                key = (item, v)
                if key not in alloc_dict:
                    alloc_dict[key] = {m: 0.0 for m in months}
                alloc_dict[key][label] += alloc

            carry = remaining  # whatever is left rolls into next month

        # After all months, if carry remains it is permanently unallocated
        if carry > 0.1:
            warnings.append(
                f"{item}: {carry:.0f} units permanently UNALLOCATED "
                f"(capacity exhausted across all months)"
            )

    # Build final allocation df
    final_rows = []
    for (item, vendor), md in alloc_dict.items():
        match = demand_df[demand_df[item_col].astype(str).str.strip() == item]
        if match.empty:
            continue
        r = {item_col: item, "Allocated Vendor": vendor}
        for lbl in months:
            r[lbl] = int(md.get(lbl, 0))
        final_rows.append(r)
    final_df = pd.DataFrame(final_rows) if final_rows else pd.DataFrame()

    # Build summary df.
    # "Rolled to Next"    = qty deferred to next month (NOT a final loss, avoid double-count)
    # "Perm Unallocated"  = qty unallocated after the LAST month  (true loss, counted once)
    summ_rows = []
    last_month = months[-1] if months else None

    for _, row in demand_df.iterrows():
        item = str(row[item_col]).strip()
        if not item or item == "nan":
            continue
        carry_s = 0.0
        for label in months:
            dcol = demand_month_map.get(label)
            d_val = 0.0
            if dcol:
                raw = row.get(dcol, 0)
                d_val = float(raw) if not pd.isna(raw) else 0.0
            own_demand   = d_val
            total_demand = carry_s + own_demand
            alloc_qty = sum(
                alloc_dict.get((item, v), {}).get(label, 0) for v in cap_dict
            )
            gap      = max(0.0, total_demand - alloc_qty)
            is_last  = (label == last_month)

            if total_demand > 0:
                summ_rows.append({
                    "Component":        item,
                    "Month":            label,
                    "Rollover In":      int(carry_s),
                    "Own Demand":       int(own_demand),
                    "Total Demand":     int(total_demand),
                    "Allocated":        int(alloc_qty),
                    "Rolled to Next":   0 if is_last else int(gap),
                    "Perm Unallocated": int(gap) if is_last else 0,
                })
            carry_s = gap
    summary_df = pd.DataFrame(summ_rows) if summ_rows else pd.DataFrame()
    return final_df, summary_df, warnings


class POAllocator(tk.Tk):

    STEPS = ["Load File", "Demand Config", "Capacity Config", "Run & Export"]

    def __init__(self):
        super().__init__()
        self.title("Smart Vendor Order Allocation Based on Capacity  v1.1")
        self.geometry("1200x840")
        self.minsize(1000, 700)
        self.configure(bg=C["bg"])
        self.file_path  = None
        self.demand_df  = None
        self.cap_df     = None
        self.final_df   = None
        self.summary_df = None
        self.cur_step   = 0
        self._build_ui()
        self.go_step(0)
        # Global mouse-wheel scroll: forward to whichever widget is under the cursor
        self.bind_all("<MouseWheel>", self._global_scroll)
        self.bind_all("<Button-4>",   self._global_scroll)
        self.bind_all("<Button-5>",   self._global_scroll)

    def _global_scroll(self, e):
        # Let the event reach the widget that is actually under the mouse
        w = e.widget
        # Walk up until we find something scrollable
        while w:
            cls = w.winfo_class()
            if cls in ("Listbox", "Canvas", "Text"):
                if e.num == 4:
                    w.yview_scroll(-1, "units")
                elif e.num == 5:
                    w.yview_scroll(1, "units")
                else:
                    w.yview_scroll(int(-e.delta / 120), "units")
                return
            if cls == "Treeview":
                if e.num == 4:
                    w.yview_scroll(-1, "units")
                elif e.num == 5:
                    w.yview_scroll(1, "units")
                else:
                    w.yview_scroll(int(-e.delta / 120), "units")
                return
            try:
                w = w.master
            except Exception:
                break

    def _build_ui(self):
        tbar = tk.Frame(self, bg=C["accent"])
        tbar.pack(fill="x")
        tk.Label(tbar, text="  Smart Vendor Order Allocation Based on Capacity  v1.1",
                 bg=C["accent"], fg="#FFFFFF",
                 font=fnt(13, bold=True)).pack(side="left", padx=16, pady=10)
        tk.Label(tbar, text="Vendor Capacity  |  Rollover Enabled",
                 bg=C["accent"], fg="#99F6E4", font=fnt(10)).pack(side="left")

        sb_wrap = tk.Frame(self, bg=C["panel"], pady=10)
        sb_wrap.pack(fill="x")
        inner_sb = tk.Frame(sb_wrap, bg=C["panel"])
        inner_sb.pack(padx=24)
        self.stepbar = StepBar(inner_sb, self.STEPS)
        self.stepbar.pack()

        tk.Frame(self, bg=C["border"], height=1).pack(fill="x")

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True)

        self.frames = {}
        for cls, key in [
            (self._build_step1, "step1"),
            (self._build_step2, "step2"),
            (self._build_step3, "step3"),
            (self._build_step4, "step4"),
        ]:
            f = ScrollFrame(body, bg=C["bg"])
            cls(f.inner)
            self.frames[key] = f

        nav = tk.Frame(self, bg=C["panel"], pady=10)
        nav.pack(fill="x", side="bottom")
        tk.Frame(nav, bg=C["border"], height=1).pack(fill="x", side="top")
        self.btn_back = make_btn(nav, "  Back",   self.go_prev, kind="ghost",   width=14)
        self.btn_next = make_btn(nav, "Next  >",  self.go_next, kind="primary",  width=14)
        self.btn_back.pack(side="left",  padx=24, pady=8)
        self.btn_next.pack(side="right", padx=24, pady=8)

    def _build_step1(self, parent):
        parent.configure(bg=C["bg"])
        tk.Label(parent, text="Step 1 - Load Excel File",
                 bg=C["bg"], fg=C["text"], font=fnt(15, bold=True)).pack(
            anchor="w", padx=28, pady=(20, 2))
        tk.Label(parent, text="Select your workbook. It must contain a Demand sheet and a Capacity sheet.",
                 bg=C["bg"], fg=C["dim"], font=fnt(10)).pack(anchor="w", padx=28, pady=(0, 12))

        dz = card(parent)
        dz.pack(fill="x", padx=28, pady=6)
        inner = tk.Frame(dz, bg=C["card"])
        inner.pack(pady=20, padx=32)
        tk.Label(inner, text="Click Browse to select an Excel file",
                 bg=C["card"], fg=C["text"], font=fnt(11, bold=True)).pack(pady=(4, 2))
        tk.Label(inner, text="Supports .xlsx and .xls",
                 bg=C["card"], fg=C["dim"], font=fnt(9)).pack()
        self.lbl_file = tk.Label(inner, text="No file selected",
                                 bg=C["card"], fg=C["dim"], font=fnt(9, italic=True))
        self.lbl_file.pack(pady=(10, 0))
        make_btn(inner, "Browse File", self._browse_file, kind="primary", width=20).pack(pady=12)

        sc = card(parent)
        sc.pack(fill="x", padx=28, pady=6)
        slabel(sc, "Sheet Selection")
        for lbl, attr in [
            ("Demand Sheet  (items and monthly demand quantities):", "cb_demand_sheet"),
            ("Capacity Sheet  (vendor names, capacities, priority):", "cb_cap_sheet"),
        ]:
            row = tk.Frame(sc, bg=C["card"])
            row.pack(fill="x", padx=16, pady=5)
            tk.Label(row, text=lbl, bg=C["card"], fg=C["text"],
                     font=fnt(10), width=48, anchor="w").pack(side="left")
            var = tk.StringVar()
            cb = make_combo(row, textvariable=var, width=30)
            cb.pack(side="left", padx=8)
            setattr(self, attr, cb)
            setattr(self, attr.replace("cb_", "var_"), var)
        make_btn(sc, "Load Sheets", self._load_sheets,
                 kind="success", width=18).pack(anchor="w", padx=16, pady=(8, 16))

    def _build_step2(self, parent):
        parent.configure(bg=C["bg"])
        tk.Label(parent, text="Step 2 - Demand Sheet Configuration",
                 bg=C["bg"], fg=C["text"], font=fnt(15, bold=True)).pack(
            anchor="w", padx=28, pady=(20, 2))
        tk.Label(parent,
                 text="Specify the item column, choose the monthly demand columns "
                      "(drag to set priority), then select your vendor columns.",
                 bg=C["bg"], fg=C["dim"], font=fnt(10),
                 wraplength=900, justify="left").pack(anchor="w", padx=28, pady=(0, 12))

        ic = card(parent)
        ic.pack(fill="x", padx=28, pady=6)
        slabel(ic, "Item / Component Column")
        hlabel(ic, "The column that uniquely identifies each product or component.")
        row = tk.Frame(ic, bg=C["card"])
        row.pack(fill="x", padx=16, pady=(0, 14))
        tk.Label(row, text="Column:", bg=C["card"], fg=C["text"],
                 font=fnt(10), width=10, anchor="w").pack(side="left")
        self.var_item_col = tk.StringVar()
        self.cb_item_col = make_combo(row, textvariable=self.var_item_col, width=32)
        self.cb_item_col.pack(side="left", padx=8)

        mc = card(parent)
        mc.pack(fill="x", padx=28, pady=6)
        slabel(mc, "Monthly Demand Columns")
        hlabel(mc, "Select one column per month. Drag to reorder - top = filled first. "
                   "Unallocated qty automatically rolls into the next month.")
        mb = tk.Frame(mc, bg=C["card"])
        mb.pack(fill="x", padx=16, pady=(4, 14))
        self.month_chooser = ColChooser(mb, title_r="Demand Months (ordered)",
                                        multi=True, height=6)
        self.month_chooser.pack(fill="x")

        vc = card(parent)
        vc.pack(fill="x", padx=28, pady=6)
        slabel(vc, "Vendor Columns")
        hlabel(vc, "Select all columns containing vendor names (e.g. Vendor 1, Vendor 2). "
                   "Use the prefix filter to quickly find matching columns.")
        pr = tk.Frame(vc, bg=C["card"])
        pr.pack(fill="x", padx=16, pady=(0, 4))
        tk.Label(pr, text="Column name prefix:", bg=C["card"],
                 fg=C["text"], font=fnt(10)).pack(side="left")
        self.var_vendor_prefix = tk.StringVar(value="Vendor")
        tk.Entry(pr, textvariable=self.var_vendor_prefix, width=16,
            bg=C["entry"], fg=C["text"], insertbackground=C["text"],
            relief="flat", font=fnt(10),
            highlightthickness=1, highlightbackground=C["border"],
            highlightcolor=C["accent"]).pack(side="left", padx=8)
        make_btn(pr, "Apply Filter",
                 lambda: self.vendor_chooser.set_prefix_filter(
                     self.var_vendor_prefix.get()),
                 kind="ghost", width=14).pack(side="left")
        vb = tk.Frame(vc, bg=C["card"])
        vb.pack(fill="x", padx=16, pady=(4, 14))
        self.vendor_chooser = ColChooser(vb, title_r="Selected Vendor Columns",
                                         multi=True, height=5)
        self.vendor_chooser.pack(fill="x")

    def _build_step3(self, parent):
        parent.configure(bg=C["bg"])
        tk.Label(parent, text="Step 3 - Capacity Sheet Configuration",
                 bg=C["bg"], fg=C["text"], font=fnt(15, bold=True)).pack(
            anchor="w", padx=28, pady=(20, 2))
        tk.Label(parent,
                 text="Map the vendor name column, the capacity per month, "
                      "and the optional priority column from your capacity sheet.",
                 bg=C["bg"], fg=C["dim"], font=fnt(10),
                 wraplength=900, justify="left").pack(anchor="w", padx=28, pady=(0, 12))

        nc = card(parent)
        nc.pack(fill="x", padx=28, pady=6)
        slabel(nc, "Vendor Name Column")
        hlabel(nc, "Column in the capacity sheet that contains vendor names.")
        row = tk.Frame(nc, bg=C["card"])
        row.pack(fill="x", padx=16, pady=(0, 14))
        tk.Label(row, text="Column:", bg=C["card"], fg=C["text"],
                 font=fnt(10), width=14, anchor="w").pack(side="left")
        self.var_vname_col = tk.StringVar()
        self.cb_vname_col = make_combo(row, textvariable=self.var_vname_col, width=30)
        self.cb_vname_col.pack(side="left", padx=8)

        pc = card(parent)
        pc.pack(fill="x", padx=28, pady=6)
        slabel(pc, "Priority Column  (optional)")
        hlabel(pc, "Lower number = filled first. Leave as (none) for equal priority.")
        row2 = tk.Frame(pc, bg=C["card"])
        row2.pack(fill="x", padx=16, pady=(0, 14))
        tk.Label(row2, text="Column:", bg=C["card"], fg=C["text"],
                 font=fnt(10), width=14, anchor="w").pack(side="left")
        self.var_prio_col = tk.StringVar(value="(none)")
        self.cb_prio_col = make_combo(row2, textvariable=self.var_prio_col, width=30)
        self.cb_prio_col.pack(side="left", padx=8)

        mcc = card(parent)
        mcc.pack(fill="x", padx=28, pady=6)
        slabel(mcc, "Month  ->  Capacity Column Mapping")
        hlabel(mcc, "For each month label (must match exactly what you set in Step 2), "
                    "select the corresponding capacity column from the capacity sheet.")
        self.month_cap_table = MonthCapTable(mcc)
        self.month_cap_table.pack(fill="x", padx=16, pady=(4, 4))
        make_btn(mcc, "Auto-populate from Step 2", self._auto_populate_months,
                 kind="ghost", width=30).pack(anchor="w", padx=16, pady=(0, 14))

    def _build_step4(self, parent):
        parent.configure(bg=C["bg"])
        tk.Label(parent, text="Step 4 - Run Allocation & Export",
                 bg=C["bg"], fg=C["text"], font=fnt(15, bold=True)).pack(
            anchor="w", padx=28, pady=(20, 2))
        tk.Label(parent,
                 text="Run the allocation. Unallocated qty from each month is automatically "
                      "rolled into the next month before that month's own demand is filled.",
                 bg=C["bg"], fg=C["dim"], font=fnt(10),
                 wraplength=900).pack(anchor="w", padx=28, pady=(0, 12))

        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=28, pady=(0, 8))
        make_btn(top, "Run Allocation",  self._run_alloc, kind="success", width=22).pack(
            side="left", padx=(0, 10))
        make_btn(top, "Export to Excel", self._export,    kind="primary", width=22).pack(
            side="left")

        sc = card(parent)
        sc.pack(fill="x", padx=28, pady=(0, 6))
        si = tk.Frame(sc, bg=C["card"])
        si.pack(fill="x", padx=16, pady=10)
        tk.Label(si, text="Status:", bg=C["card"], fg=C["dim"],
                 font=fnt(10)).pack(side="left")
        self.lbl_status = tk.Label(si, text="Ready to run",
                                   bg=C["card"], fg=C["success"], font=fnt(10, bold=True))
        self.lbl_status.pack(side="left", padx=8)
        self.prog_bar = ttk.Progressbar(si, mode="indeterminate", length=200)
        self.prog_bar.pack(side="right")

        self.warn_frame = tk.Frame(parent, bg=C["bg"])
        self.warn_frame.pack(fill="x", padx=28, pady=(0, 4))

        nb_wrap = card(parent)
        nb_wrap.pack(fill="both", expand=True, padx=28, pady=(0, 12))
        st = ttk.Style()
        st.configure("Pro.TNotebook",     background=C["card"], borderwidth=0)
        st.configure("Pro.TNotebook.Tab",
            background=C["card2"], foreground=C["dim"],
            padding=(18, 7), font=(FF, 10))
        st.map("Pro.TNotebook.Tab",
            background=[("selected", C["panel"])],
            foreground=[("selected", C["accent"])])
        self.nb = ttk.Notebook(nb_wrap, style="Pro.TNotebook")
        self.nb.pack(fill="both", expand=True, padx=2, pady=2)
        self.result_table  = ResultsTable(self.nb)
        self.summary_table = ResultsTable(self.nb)
        self.nb.add(self.result_table,  text="  Allocation Results  ")
        self.nb.add(self.summary_table, text="  Summary with Rollover  ")

    def go_step(self, idx):
        keys = ["step1", "step2", "step3", "step4"]
        for f in self.frames.values():
            f.pack_forget()
        self.frames[keys[idx]].pack(fill="both", expand=True)
        self.stepbar.set_step(idx)
        self.cur_step = idx
        self.btn_back.config(state="normal" if idx > 0 else "disabled")
        self.btn_next.config(text="Finish" if idx == 3 else "Next  >")

    def go_next(self):
        try:
            if not self._validate(self.cur_step):
                return
            if self.cur_step < 3:
                self._on_advance(self.cur_step)
                self.go_step(self.cur_step + 1)
        except Exception as ex:
            messagebox.showerror("Navigation Error", str(ex))

    def go_prev(self):
        if self.cur_step > 0:
            self.go_step(self.cur_step - 1)

    def _validate(self, step):
        if step == 0 and (self.demand_df is None or self.cap_df is None):
            messagebox.showwarning("Smart Vendor Order Allocation", "Please select a file and click Load Sheets first.")
            return False
        if step == 1:
            if not self.var_item_col.get():
                messagebox.showwarning("Smart Vendor Order Allocation", "Please select an Item column.")
                return False
            if not self.month_chooser.get_selected():
                messagebox.showwarning("Smart Vendor Order Allocation", "Please select at least one demand month column.")
                return False
            if not self.vendor_chooser.get_selected():
                messagebox.showwarning("Smart Vendor Order Allocation", "Please select at least one Vendor column.")
                return False
        if step == 2:
            if not self.var_vname_col.get():
                messagebox.showwarning("Smart Vendor Order Allocation", "Please select the Vendor Name column.")
                return False
            if not self.month_cap_table.get_mapping():
                messagebox.showwarning("Smart Vendor Order Allocation", "Please add at least one Month -> Capacity mapping.")
                return False
        return True

    def _on_advance(self, from_step):
        if from_step == 0:
            cols = list(self.demand_df.columns)
            self.cb_item_col["values"] = cols
            self.month_chooser.set_columns(cols)
            self.vendor_chooser.set_columns(cols)
            self.vendor_chooser.set_prefix_filter(self.var_vendor_prefix.get())
        if from_step == 1:
            cols = list(self.cap_df.columns)
            self.cb_vname_col["values"] = cols
            self.cb_prio_col["values"]  = ["(none)"] + cols
            self.month_cap_table.set_columns(cols)

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not path:
            return
        self.file_path = path
        self.lbl_file.config(text=os.path.basename(path), fg=C["success"])
        try:
            sheets = pd.ExcelFile(path).sheet_names
            self.cb_demand_sheet["values"] = sheets
            self.cb_cap_sheet["values"]    = sheets
            if sheets:
                self.cb_demand_sheet.set(sheets[0])
                self.cb_cap_sheet.set(sheets[-1] if len(sheets) > 1 else sheets[0])
                self._load_sheets()
            Toast.show(self, f"{len(sheets)} sheets found", "success")
        except Exception as ex:
            messagebox.showerror("Load Error", str(ex))

    def _load_sheets(self):
        if not self.file_path:
            messagebox.showwarning("Smart Vendor Order Allocation", "Please select a file first.")
            return
        ds = self.var_demand_sheet.get()
        cs = self.var_cap_sheet.get()
        if not ds or not cs:
            messagebox.showwarning("Smart Vendor Order Allocation", "Please select both sheets.")
            return
        try:
            self.demand_df = pd.read_excel(self.file_path, sheet_name=ds)
            self.cap_df    = pd.read_excel(self.file_path, sheet_name=cs)
            self.demand_df.columns = self.demand_df.columns.str.strip()
            self.cap_df.columns    = self.cap_df.columns.str.strip()
            Toast.show(self,
                       f"Loaded  {len(self.demand_df)} demand rows, "
                       f"{len(self.cap_df)} capacity rows", "success")
        except Exception as ex:
            messagebox.showerror("Load Error", str(ex))

    def _auto_populate_months(self):
        months = self.month_chooser.get_selected()
        self.month_cap_table.clear()
        cap_cols = list(self.cap_df.columns) if self.cap_df is not None else []
        for m in months:
            match = next((c for c in cap_cols if m.lower() in str(c).lower()), "")
            self.month_cap_table.add_row(month=m, col=match)

    def _run_alloc(self):
        if not self._validate(2):
            return
        self.lbl_status.config(text="Running allocation...", fg=C["warn"])
        self.prog_bar.start(10)
        for w in self.warn_frame.winfo_children():
            w.destroy()

        def worker():
            try:
                pc = self.var_prio_col.get()
                if pc == "(none)":
                    pc = None
                final, summary, warns = run_allocation(
                    demand_df         = self.demand_df,
                    cap_df            = self.cap_df,
                    item_col          = self.var_item_col.get(),
                    month_cols        = self.month_chooser.get_selected(),
                    vendor_cols       = self.vendor_chooser.get_selected(),
                    vendor_name_col   = self.var_vname_col.get(),
                    month_cap_mapping = self.month_cap_table.get_mapping(),
                    priority_col      = pc,
                )
                self.final_df   = final
                self.summary_df = summary
                self.after(0, lambda: self._show_results(final, summary, warns))
            except Exception as ex:
                self.after(0, lambda: self._alloc_error(str(ex)))

        threading.Thread(target=worker, daemon=True).start()

    def _show_results(self, final, summary, warns):
        self.prog_bar.stop()
        self.lbl_status.config(
            text=f"Done  -  {len(final)} allocation rows, {len(summary)} summary rows",
            fg=C["success"])
        for w in self.warn_frame.winfo_children():
            w.destroy()
        if warns:
            wc = card(self.warn_frame)
            wc.pack(fill="x", pady=4)
            hf = tk.Frame(wc, bg=C["wbg"])
            hf.pack(fill="x")
            tk.Label(hf, text=f"  {len(warns)} Warning(s) - unallocated after all months exhausted",
                     bg=C["wbg"], fg=C["wfg"],
                     font=fnt(10, bold=True)).pack(anchor="w", padx=12, pady=6)
            for wt in warns[:10]:
                tk.Label(wc, text=f"    - {wt}",
                         bg=C["card"], fg=C["warn"], font=fnt(9)).pack(anchor="w", padx=12)
            if len(warns) > 10:
                tk.Label(wc, text=f"    ... and {len(warns) - 10} more",
                         bg=C["card"], fg=C["dim"],
                         font=fnt(9)).pack(anchor="w", padx=12, pady=(0, 6))
        if not final.empty:
            self.result_table.load(final)
        if not summary.empty:
            self.summary_table.load(summary)
        self.nb.select(0)
        Toast.show(self, "Allocation complete", "success")

    def _alloc_error(self, msg):
        self.prog_bar.stop()
        self.lbl_status.config(text="Error during allocation", fg=C["error"])
        messagebox.showerror("Allocation Error", msg)

    def _export(self):
        if self.final_df is None or self.final_df.empty:
            messagebox.showwarning("Smart Vendor Order Allocation", "Run the allocation first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save Output", defaultextension=".xlsx",
            initialfile="SmartVendorOrderAllocation_Output.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("CSV file", "*.csv")])
        if not path:
            return
        if path.endswith(".csv"):
            self.final_df.to_csv(path, index=False)
            Toast.show(self, "CSV exported", "success")
            return
        self._export_xlsx(path)

    def _export_xlsx(self, path):
        try:
            wb = Workbook()
            first = True
            for sname, df in [("Allocation", self.final_df), ("Summary", self.summary_df)]:
                if df is None or df.empty:
                    continue
                ws = wb.active if first else wb.create_sheet(sname)
                if first:
                    ws.title = sname
                    first = False

                hdr_fill = PatternFill("solid", fgColor="0D9488")
                hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                thin     = Side(style="thin", color="D1D5DB")
                brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
                even_f   = PatternFill("solid", fgColor="F9FAFB")
                num_f    = PatternFill("solid", fgColor="CCFBF1")
                roll_f   = PatternFill("solid", fgColor="F0FDFA")
                ctr      = Alignment(horizontal="center", vertical="center")

                cols = list(df.columns)
                roll_idx = (cols.index("Rolled to Next") + 1 if "Rolled to Next" in cols
                            else (cols.index("Rollover In") + 1 if "Rollover In" in cols else None))

                for ci, c in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=ci, value=c)
                    cell.fill = hdr_fill; cell.font = hdr_font
                    cell.alignment = ctr; cell.border = brd
                    ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(c)) + 6)

                for ri, (_, row) in enumerate(df.iterrows(), 2):
                    has_roll = False
                    if roll_idx:
                        rv = row.iloc[roll_idx - 1]
                        has_roll = isinstance(rv, (int, float)) and rv > 0
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        cell.border = brd; cell.alignment = ctr
                        cell.font = Font(name="Calibri", size=10)
                        if has_roll:
                            cell.fill = roll_f
                        elif ri % 2 == 0:
                            cell.fill = even_f
                        if isinstance(val, (int, float)) and val > 0 and not has_roll:
                            cell.fill = num_f

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

            wb.save(path)
            Toast.show(self, f"Saved: {os.path.basename(path)}", "success")
        except Exception as ex:
            messagebox.showerror("Export Error", str(ex))


if __name__ == "__main__":
    app = POAllocator()
    app.mainloop()