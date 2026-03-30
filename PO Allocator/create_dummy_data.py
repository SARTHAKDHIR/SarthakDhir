"""
Creates a realistic dummy input Excel file for AlloFlow (bracelet product line).
Two sheets: Demand and Capacity
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Demand Sheet ─────────────────────────────────────────────────────────────
# Bracelet components / SKUs
demand_data = {
    "Component": [
        "Charm Bracelet - Gold 18K",
        "Charm Bracelet - Silver 925",
        "Bangle - Rose Gold 14K",
        "Bangle - Stainless Steel",
        "Tennis Bracelet - Diamond",
        "Tennis Bracelet - Cubic Zirconia",
        "Cuff Bracelet - Brass",
        "Cuff Bracelet - Sterling Silver",
        "Beaded Bracelet - Lapis Lazuli",
        "Beaded Bracelet - Tiger Eye",
        "Link Bracelet - Gold 10K",
        "Link Bracelet - Silver",
        "Leather Bracelet - Brown",
        "Leather Bracelet - Black",
        "Chain Bracelet - Box Chain Gold",
    ],
    "Jan-25": [1200, 900,  800,  2000, 300, 1500, 700,  600,  2500, 2200, 450,  850,  1800, 1600, 500 ],
    "Feb-25": [1100, 850,  750,  1800, 280, 1400, 650,  580,  2300, 2000, 420,  800,  1700, 1500, 480 ],
    "Mar-25": [1400, 1000, 900,  2200, 350, 1700, 800,  700,  2800, 2400, 500,  950,  2000, 1800, 560 ],
    "Apr-25": [1300, 950,  850,  2100, 320, 1600, 750,  650,  2600, 2300, 480,  900,  1900, 1700, 530 ],
    "May-25": [1500, 1100, 1000, 2400, 380, 1800, 870,  760,  3000, 2600, 540,  1000, 2100, 1900, 590 ],
    "Jun-25": [1600, 1200, 1050, 2600, 400, 1950, 900,  800,  3200, 2800, 570,  1050, 2200, 2000, 620 ],
    "Vendor 1": [
        "GoldCraft Co.", "GoldCraft Co.", "RoseGold Works", "MetalMasters",
        "DiamondLine", "DiamondLine", "BrassPro", "SilverSmith Ltd",
        "GemBeads Inc", "GemBeads Inc", "GoldCraft Co.", "SilverSmith Ltd",
        "LeatherArt", "LeatherArt", "GoldCraft Co.",
    ],
    "Vendor 2": [
        "SilverSmith Ltd", "MetalMasters", "MetalMasters", "BrassPro",
        "GoldCraft Co.", "SilverSmith Ltd", "MetalMasters", "BrassPro",
        "EastBeads Co.", "EastBeads Co.", "SilverSmith Ltd", "GoldCraft Co.",
        "MetalMasters", "MetalMasters", "SilverSmith Ltd",
    ],
    "Vendor 3": [
        "MetalMasters", "BrassPro", "GoldCraft Co.", "SilverSmith Ltd",
        "SilverSmith Ltd", "MetalMasters", "GoldCraft Co.", "MetalMasters",
        "SilverSmith Ltd", "BrassPro", "", "", "BrassPro", "BrassPro", "MetalMasters",
    ],
}

demand_df = pd.DataFrame(demand_data)

# ── Capacity Sheet ────────────────────────────────────────────────────────────
capacity_data = {
    "Vendor Name": [
        "GoldCraft Co.",
        "SilverSmith Ltd",
        "RoseGold Works",
        "MetalMasters",
        "DiamondLine",
        "BrassPro",
        "GemBeads Inc",
        "EastBeads Co.",
        "LeatherArt",
    ],
    "Priority": [1, 2, 1, 3, 1, 4, 2, 3, 1],
    "Cap Jan-25": [2500, 2000, 1200, 3500, 600,  1800, 4000, 3000, 3500],
    "Cap Feb-25": [2300, 1900, 1100, 3200, 560,  1700, 3800, 2800, 3200],
    "Cap Mar-25": [2800, 2200, 1300, 4000, 700,  2000, 4500, 3300, 4000],
    "Cap Apr-25": [2600, 2100, 1250, 3800, 640,  1900, 4200, 3100, 3800],
    "Cap May-25": [3000, 2400, 1400, 4200, 760,  2100, 4800, 3500, 4200],
    "Cap Jun-25": [3200, 2600, 1500, 4500, 800,  2200, 5000, 3700, 4500],
}

capacity_df = pd.DataFrame(capacity_data)

# ── Write to Excel ────────────────────────────────────────────────────────────
output_path = r"c:\Users\sarth\OneDrive\Desktop\Git Hub project Claude\PO Allocator\input_bracelet_dummy.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    demand_df.to_excel(writer,  sheet_name="Demand",   index=False)
    capacity_df.to_excel(writer, sheet_name="Capacity", index=False)

# ── Style the workbook ────────────────────────────────────────────────────────
wb = load_workbook(output_path)

teal_fill  = PatternFill("solid", fgColor="0D9488")
teal_light = PatternFill("solid", fgColor="CCFBF1")
even_fill  = PatternFill("solid", fgColor="F0FDFA")
hdr_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
data_font  = Font(name="Calibri", size=10)
ctr        = Alignment(horizontal="center", vertical="center", wrap_text=False)
left       = Alignment(horizontal="left",   vertical="center")
thin       = Side(style="thin", color="D1D5DB")
brd        = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    # Header row
    for cell in ws[1]:
        cell.fill      = teal_fill
        cell.font      = hdr_font
        cell.alignment = ctr
        cell.border    = brd

    # Data rows
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = even_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in row:
            cell.font      = data_font
            cell.fill      = fill
            cell.border    = brd
            # Numbers centered, text left-aligned
            if isinstance(cell.value, (int, float)):
                cell.alignment = ctr
            else:
                cell.alignment = left

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(14, max_len + 4)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

wb.save(output_path)
print(f"Done! Saved to: {output_path}")
print(f"\nDemand sheet  : {len(demand_df)} rows x {len(demand_df.columns)} columns")
print(f"Capacity sheet: {len(capacity_df)} rows x {len(capacity_df.columns)} columns")
